from pathlib import Path

from openpyxl import load_workbook

# Hardcoded path to the Excel workbook for Prototype 001.
WORKBOOK_PATH = Path(__file__).resolve().parent.parent / "data" / "Final_Budget.xlsx"

def open_workbook():
    """Open the hardcoded workbook and return its name and worksheet names."""
    try:
        workbook = load_workbook(WORKBOOK_PATH, read_only=True)
        worksheet_names = workbook.sheetnames
        workbook_name = WORKBOOK_PATH.name
        workbook.close()
        return workbook_name, worksheet_names
    except FileNotFoundError:
        print(f"Error: The workbook '{WORKBOOK_PATH}' was not found.")
        return None, None



def sheet_count():
    """Return the number of worksheets in the workbook."""
    _, worksheet_names = open_workbook()
    return len(worksheet_names)


def print_workbook_info():
    """Print the name of the workbook and all worksheet names."""
    workbook_name, worksheet_names = open_workbook()
    print(f"Workbook: {workbook_name}")
    print("\nWorksheets:")
    for i, sheet_name in enumerate(worksheet_names, start=1):
        print(f"{i}. {sheet_name}")

try:
    print_workbook_info()
except FileNotFoundError:
    print(f"Error: The workbook '{WORKBOOK_PATH}' was not found.")

if __name__ == "__main__":
    print_workbook_info()

def validate_workbook():
    """Validate that the workbook exists and can be opened."""
    try:
        load_workbook(WORKBOOK_PATH, read_only=True)
        return True
    except FileNotFoundError:
        print(f"Error: The workbook '{WORKBOOK_PATH}' was not found.")
        return False




def read_cell(sheet_name, cell_reference):
    """Read the value of a cell from a worksheet."""
    try:
        workbook = load_workbook(WORKBOOK_PATH, read_only=True)
    except FileNotFoundError:
        print(f"Error: The workbook '{WORKBOOK_PATH}' was not found.")
        return None
 

    try:
        worksheet = workbook[sheet_name]
    except KeyError:
        print(f"Error: Worksheet '{sheet_name}' does not exist in '{WORKBOOK_PATH.name}'.")
        workbook.close()
        return None
    print("Worksheet:", worksheet.title)
    print("Cell:", cell_reference)
    print("Value:", worksheet[cell_reference].value)

    cell_value = worksheet[cell_reference].value
    workbook.close()
    return cell_value


if __name__ == "__main__":
    value = read_cell("Sheet1", "A1")
    print(value)

def write_cell(sheet_name, cell_reference, value):
    """Write a value to a worksheet cell."""
    try:
        workbook = load_workbook(WORKBOOK_PATH)
    except FileNotFoundError:
        print(f"Error: The workbook '{WORKBOOK_PATH}' was not found.")
        return None

    try:
        try:
            worksheet = workbook[sheet_name]
        except KeyError:
            print(f"Error: Worksheet '{sheet_name}' does not exist in '{WORKBOOK_PATH.name}'.")
            return None

        # Assign the value to the cell
        try:
            worksheet[cell_reference].value = value
        except Exception as e:
            print(f"Error writing to cell '{cell_reference}' in '{sheet_name}': {e}")
            return None

        # Save changes and return the written value
        try:
            workbook.save(WORKBOOK_PATH)
        except Exception as e:
            print(f"Error saving workbook '{WORKBOOK_PATH}': {e}")
            return None

        return value
    finally:
        workbook.close()

print(WORKBOOK_PATH)